import openpyxl
import os
import xlsxwriter
#新建excel
path='C:\\Users\\Administrator\\Desktop\\1111\\'
wb2=openpyxl.Workbook()
wb2.save(path+'test.xlsx')
print('新建成功')

#读取数据
# workbook = xlsxwriter.Workbook(path+'test.xlsx')  # 创建一个excel文件
def show_filess(filename,path1):
    sheet1name = 'new630'
    sheet2name = 'new930'
    sheet3name = 'new1031'
    sheet4name = 'new1231'
    sheetname = [sheet1name, sheet2name, sheet3name, sheet4name]

    # worksheet.set_column('A:A', 20)  # 设置第一列宽度为20像素
    # bold = workbook.add_format({'bold': True})  # 设置一个加粗的格式对象
    # worksheet.write('A1', 'HELLO')  # 在A1单元格写上HELLO
    # worksheet.write('A2', 'WORLD', bold)  # 在A2上写上WORLD,并且设置为加粗
    # worksheet.write('B2', U'中文测试', bold)  # 在B2上写上中文加粗
    # worksheet.write(2, 0, 32)  # 使用行列的方式写上数字32,35,5
    # worksheet.write(3, 0, 35.5)  # 使用行列的时候第一行起始为0,所以2,0代表着第三行的第一列,等价于A4
    # worksheet.write(4, 0, '=SUM(A3:A4)')  # 写上excel公式
    # workbook.close()

    wb1 = openpyxl.load_workbook(path1+filename)
    wb2 = openpyxl.load_workbook(path+'test.xlsx')
    # sheets1=wb1.get_sheet_names()#获取sheet页
    # sheets2=wb2.get_sheet_names()
    for name in sheetname:
        if (name in wb2.sheetnames):
            worksheet = wb2.get_worksheet_by_name(name)  # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
        else:
            worksheet = wb2.add_worksheet(name)  # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
            sheet1 = wb1.get_sheet_by_name(name)
            sheet2 = wb2.get_sheet_by_name(name)
            max_row = sheet1.max_row  # 最大行数
            max_column = sheet1.max_column  # 最大列数

            for m in range(1, max_row + 1):
                for n in range(97, 97 + max_column):  # chr(97)='a'
                    n = chr(n)  # ASCII字符
                    i = '%s%d' % (n, m)  # 单元格编号
                    cell1 = sheet1[i].value  # 获取data单元格数据
                    sheet2[i].value = cell1  # 赋值到test单元格

            wb2.save('test.xlsx')  # 保存数据
            wb1.close()  # 关闭excel
            wb2.close()



def show_files(path, all_files):
    # 首先遍历当前目录所有文件及文件夹
    file_list = os.listdir(path)
    # 准备循环判断每个元素是否是文件夹还是文件，是文件的话，把名称传入list，是文件夹的话，递归
    for file in file_list:
        # 利用os.path.join()方法取得路径全名，并存入cur_path变量，否则每次只能遍历一层目录
        cur_path = os.path.join(path, file)
        # 判断是否是文件夹
        if os.path.isdir(cur_path):
            show_files(cur_path, all_files)
        else:
            all_files.append(file)

    return all_files


# 传入空的list接收文件名
path1='C:\\Users\\Administrator\\Desktop\\新建文件夹\\所有组件各阶段数据统计\\所有组件各阶段数据统计\\'
contents = show_files(path1, [])
# 循环打印show_files函数返回的文件名列表
for content in contents:
    print(content)
    show_filess(content,path1)
