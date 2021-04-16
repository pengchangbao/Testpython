import openpyxl
import os
import xlsxwriter
#新建excel
path='C:\\Users\\Administrator\\Desktop\\1111\\'
wb2=openpyxl.Workbook()
sheet1name = 'new630'
sheet2name = 'new930'
sheet3name = 'new1031'
sheet4name = 'new1231'
sheetname = [sheet1name, sheet2name, sheet3name, sheet4name]
for index ,name in enumerate(sheetname):
    wb2.create_sheet(name,index)
print('新建成功')
wb2.save(path+'test.xlsx')

#读取数据
# workbook = xlsxwriter.Workbook(path+'test.xlsx')  # 创建一个excel文件
def show_filess(filename,path1):

    wb1 = openpyxl.load_workbook(path1+filename)
    for name in sheetname:
        print('name:'+name)
        sheet1 = wb1.get_sheet_by_name(name)
        sheet2 = wb2.get_sheet_by_name(name)
        max_row = sheet1.max_row  # 最大行数
        max_column = sheet1.max_column  # 最大列数

        for m in range(1, max_row + 1):
            for n in range(97, 97 + max_column):  # chr(97)='a'
                n = chr(n)  # ASCII字符
                i = '%s%d' % (n, m)  # 单元格编号
                cell1 = sheet1[i].value  # 获取data单元格数据
                sheet2[i].value = cell1  # 赋值到test单元格

        wb1.close()  # 关闭excel




def show_files(path, all_files):
    # 首先遍历当前目录所有文件及文件夹
    file_list = os.listdir(path)
    # 准备循环判断每个元素是否是文件夹还是文件，是文件的话，把名称传入list，是文件夹的话，递归
    for file in file_list:
        # 利用os.path.join()方法取得路径全名，并存入cur_path变量，否则每次只能遍历一层目录
        cur_path = os.path.join(path, file)
        # 判断是否是文件夹
        if os.path.isdir(cur_path):
            show_files(cur_path, all_files)
        else:
            all_files.append(file)

    return all_files


# 传入空的list接收文件名
path1='C:\\Users\\Administrator\\Desktop\\新建文件夹\\所有组件各阶段数据统计\\所有组件各阶段数据统计\\'
contents = show_files(path1, [])
# 循环打印show_files函数返回的文件名列表
for content in contents:
    print(content)
    show_filess(content,path1)
    wb2.save(path + 'test.xlsx')  # 保存数据
    wb2.close()